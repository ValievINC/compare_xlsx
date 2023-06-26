import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import sys
import subprocess


def compare_pages(standard_page, student_page):
    """
    Сравниваем содержимое эталонной страницы со страницей студента.
    """
    status = False

    standard = pd.read_excel(file1, sheet_name=standard_page)
    student = pd.read_excel(file2, sheet_name=student_page)

    if student.empty:
        print(f'{{"verdict": \'WrongAnswer\', \'error\': \'Данных на странице \'{student_page}\' недостаточно \'}}')
        return status

    elif standard.shape != student.shape or not standard.equals(student):

        standard = standard.fillna("N/A")
        student = student.fillna("N/A")

        for row in range(standard.shape[0]):
            for col in range(standard.shape[1]):
                if standard.iloc[row, col] != student.iloc[row, col]:
                    cell_label = get_column_letter(col + 1) + str(row + 2)
                    print(f'{{"verdict": \'WrongAnswer\', \'error\': \'Различия на странице \'{standard_page}\'. Данные '
                          f'в ячейках не совпадают. Первая несовпадающая ячейка: \'{cell_label}\' Корректные данные '
                          f'в ячейке {cell_label}: {standard.iloc[row, col]}\'}}')
                    break
            else:
                continue
            break

        return status

    else:
        status = True

    return status


def compare_xlsx(file1, file2):
    """
    В этой функции сравниваем эталонную таблицу с таблицей студента.
    Если названия страниц в таблице совпадают, то вызываем функцию сравнения страниц.
    """
    status = False

    standard = load_workbook(filename=file1)
    student_work = load_workbook(filename=file2)

    standard_pages = standard.sheetnames
    student_pages = student_work.sheetnames

    if sorted(standard_pages) != sorted(student_pages):
        difference = [item for item in standard_pages if item not in student_pages]
        print(f'{{"verdict": \'WrongAnswer\', \'error\': \'Не хватает страниц {list(difference)}\'}}')
        return

    for standard_page, student_page in zip(standard_pages, student_pages):
        status = compare_pages(standard_page, student_page)
        if not status:
            break

    if status:
        print('{"verdict": \'Ok\'}')


def run_solution():
    """
    Запускаем код студента и ловим ошибки. Если поймали ошибку - выводим содержимое и заканчиваем работу кода.
    """
    solution = subprocess.run([sys.executable, 'solution.py'], capture_output=True, text=True)

    if solution.returncode != 0:
        error_output = solution.stderr.strip()
        formatted_error = error_output.replace('\\r\\n', '\n').replace('\\', '')
        print("Произошла ошибка при выполнении процесса:", file=sys.stderr)
        print(formatted_error, file=sys.stderr)
        sys.exit(1)


# run_solution()

file1 = "etalon.xlsx"
file2 = "report.xlsx"

compare_xlsx(file1, file2)